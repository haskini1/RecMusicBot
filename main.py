import telebot
import config
import dbworker
from telebot import types
import pandas as pd
from openpyxl import load_workbook

def group_to_estimate(group):
    return (group - 210) % 3 + 211


def newlenght(*args):
    l = 0
    for u in args:
        l = l + u ** 2
    return l ** 0.5


# скаляр векторов
def scalar(a1, b1, c1, d1, a2, b2, c2, d2):
    return a1 * a2 + b1 * b2 + c1 * c2 + d1 * d2


# длина вектора
def lenghtvector(x, y, z, u):
    return (x ** 2 + y ** 2 + z ** 2 + u ** 2) ** 0.5


# косинусное расстояние
def cosinus(a1, b1, c1, d1, a2, b2, c2, d2):
    return scalar(a1, b1, c1, d1, a2, b2, c2, d2) / ((lenghtvector(a1, b1, c1, d1)) * lenghtvector(a2, b2, c2, d2))


students = [25, 25, 24]  # количество студентов в каждой группе

#RS = pd.read_excel('D:/verygoodfile.xlsx')
RS = pd.read_excel('verygoodfile.xlsx')
# создадим двумерные массивы из нулей размером (students[index_group]+2) x (11)
table211 = [[0] * 10 for i in range(students[0])]
table212 = [[0] * 10 for i in range(students[1])]
table213 = [[0] * 10 for i in range(students[2])]
swp = [0] * 10
for i in range(10):
    swp[i] = "Singer " + "" + str(i)

for i in range(students[0]):
    for j in range(10):
        table211[i][j] = RS[swp[j]][i]

swp = [0] * 10
for i in range(10):
    swp[i] = "Singer " + "1" + str(i)

for i in range(students[1]):
    for j in range(10):
        table212[i][j] = RS[swp[j]][i]

swp = [0] * 10
for i in range(10):
    swp[i] = "Singer " + "2" + str(i)

for i in range(students[2]):
    for j in range(10):
        table213[i][j] = RS[swp[j]][i]


# for i in ('StudentName', swp[0], swp[1], swp[2], swp[3], swp[4], swp[5], swp[6], swp[7], swp[8], swp[9]):
# for j in


def rec(estimating_group, usermark0, usermark1, usermark2, usermark3):
    # создадим префиксы, т.е. первую цифру исполнителя чтобы опредеилть каких исполнителей давать на оценку
    if estimating_group == 211:
        prefix = ""
    elif estimating_group == 212:
        prefix = "1"
    elif estimating_group == 213:
        prefix = "2"

    swp = [0] * 10
    for i in range(10):
        swp[i] = "Singer " + prefix + str(i)

    RS['alpha'] = -2
    alpha_list = [0] * (students[estimating_group - 211] + 1)

    for i in range(1, students[estimating_group - 211] + 1):
        m1 = int(RS[swp[0]][i])
        m2 = int(RS[swp[1]][i])
        m3 = int(RS[swp[2]][i])
        m4 = int(RS[swp[3]][i])
        #print(i, cosinus(m1 - 3, m2 - 3, m3 - 3, m4 - 3, usermark0 - 3, usermark1 - 3, usermark2 - 3, usermark3 - 3))
        RS.loc[i, 'alpha'] = cosinus(m1 - 3, m2 - 3, m3 - 3, m4 - 3, usermark0 - 3, usermark1 - 3, usermark2 - 3,
                                     usermark3 - 3)
        alpha_list[i] = cosinus(m1 - 3, m2 - 3, m3 - 3, m4 - 3, usermark0 - 3, usermark1 - 3, usermark2 - 3,
                                usermark3 - 3)

    singer_coef = [-100] * 10

    for i in range(4, 10):
        numerator = 0
        denominator = 0
        d = 0
        for j in range(1, students[estimating_group - 211] + 1):
            numerator += RS['alpha'][j] * RS[swp[i]][j]
        singer_mark_list = [0] * (students[estimating_group - 211] + 1)
        for j in range(1, students[estimating_group - 211] + 1):
            singer_mark_list[j] = RS[swp[i]][j]
        denominator = newlenght(*singer_mark_list) * newlenght(*alpha_list)
        singer_coef[i] = 2 * numerator / denominator

    best_singers = [0] * 10
    for i in range(0, 10):
        for j in range(0, 10):
            if singer_coef[i] > singer_coef[j]:
                best_singers[i] += 1

    for i in range(0, 4):
        best_singers[i] = -100

    def indexmax(*arr):
        answer_list = [0] * 6
        for i in range(0, 6):
            answer_list[i] = best_singers.index(max(best_singers))
            best_singers[best_singers.index(max(best_singers))] = - 100
        return answer_list

    answer = []
    alpha_answer = []
    for i in indexmax(best_singers):
        #print(RS[swp[i]][0])
        answer.append(RS[swp[i]][0])
        alpha_answer.append(singer_coef[i])
    return answer, alpha_answer


"""  ********************************** bot ************************************************************ """

user_data = [0]*8
bot = telebot.TeleBot(config.API_TOKEN)


@bot.message_handler(commands=['help'])
def cmd_help(message):
    bot.send_message(message.chat.id,
                                      'Возможные команды: \n'
                                      '/help - все команды \n'
                                      '/start - начать использовать бота \n'
                                      '/reset - начать всё с начала'
                     )


@bot.message_handler(commands=['start'])
def cmd_start(message):
    bot.send_message(message.chat.id, 'Привет! Я - бот, который может порекомендовать тебе музыкальных исполнителей. Готов(а)? \n'
                                      'Напиши ДА если готов(а)')
    dbworker.set_state(message.chat.id, config.States.S_START.value)


@bot.message_handler(commands=['reset'])
def cmd_reset(message):
    bot.send_message(message.chat.id, "Начнем по-новой! Готов(а) получить рекомендации? \n"
                                      "Ответь ДА если готов(а)")
    dbworker.set_state(message.chat.id, config.States.S_START.value)


@bot.message_handler(func=lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_START.value) #првоерка CONSENT
def user_start(message):
    if message.text.upper() == "ДА":
        bot.send_message(message.chat.id,
            'Хорошо.')
        user_data[0] = message.chat.id
        user_data[1] = "YES"
        bot.send_message(message.chat.id,
                         'Наша команда хочет проверить насколько хорошо работают '
                         'наши рекомендации, поэтому мы попросим вас оценивать исполнителей честно и '
                         'проверить насколько они правдивы. Мы соберем и проанализируем эти данные, '
                         'чтобы сделать эту рекомендательную систему лучше! \n'
                         'Будешь отвечать честно или побалуешься? \n'
                         '\n'
                         '\n'
                         'Напиши БАЛОВАТЬСЯ если будешь баловаться и ЧЕСТНО если будешь отвечать честно')
        dbworker.set_state(message.chat.id, config.States.S_READY.value)
    else:
        bot.send_message(message.chat.id, 'Чтобы сбросить всё и начать пользоваться ботом введи /reset')
        dbworker.set_state(message.chat.id, config.States.S_START.value)


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_READY.value)
def user_ready(message):
    if message.text.upper() == "БАЛОВАТЬСЯ":
        user_data[2] = "NO"
        bot.send_message(message.chat.id, "Ок, твои результаты мы не будем учитывать. Можешь эксперементировать!")
        bot.send_message(message.chat.id,
                         'Из какой ты группы? 211, 212 или 213? Просто напиши номер группы'
                         '\n'
                         'Если ты не студент 1 курса БСЦ, то напиши 210'
                         )
        dbworker.set_state(message.chat.id, config.States.S_CONSENT.value)
    elif message.text.upper() == "ЧЕСТНО":
        user_data[2] = "YES"
        bot.send_message(message.chat.id, 'Ок, спасибо. Мы будем использовать твои ответы в статистике')
        bot.send_message(message.chat.id,
                         'Из какой ты группы? 211, 212 или 213? Просто напиши номер группы'
                         '\n'
                         'Если ты не студент 1 курса БСЦ, то напиши 210'
                         )
        dbworker.set_state(message.chat.id, config.States.S_CONSENT.value)

@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_CONSENT.value)
def user_choose_group(message):
    if message.text == "211":
        user_data[3] = 211
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_211_1.value)
        bot.send_message(message.chat.id, "Оцени предложенных исполнителей оценкой от 1 до 5, где \n"
                                          "1 - тебе совсем не нравится исполнитель \n"
                                          "2 - просто не нравится исполнитель \n"
                                          "3 - не слышал / равнодушен(на) \n"
                                          "4 - нравится исполнитель \n"
                                          "5 - очень нравится исполнитель \n"
                                          "\n"
                                          "Оценивай исполнителей через пробел. Все четыре оценки не должны быть одинаковыми. \n"
                                          "Пример того, что ты должен(на) ввести: \n"
                                          "4 1 2 3")
        bot.send_message(message.chat.id, 'Sam Smith, Sia, Fall Out Boy, Drake')
    elif message.text == "212":
        user_data[3] = 212
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_212_1.value)
        bot.send_message(message.chat.id, "Оцени предложенных исполнителей оценкой от 1 до 5, где \n"
                                          "1 - тебе совсем не нравится исполнитель \n"
                                          "2 - просто не нравится исполнитель \n"
                                          "3 - не слышал / равнодушен(на) \n"
                                          "4 - нравится исполнитель \n"
                                          "5 - очень нравится исполнитель \n"
                                          "\n"
                                          "Оценивай исполнителей через пробел. Все четыре оценки не должны быть одинаковыми. \n"
                                          "Пример того, что ты должен(на) ввести: \n"
                                          "4 1 2 3")
        bot.send_message(message.chat.id, 'Скриптонит, Валерий Меладзе, Кино, Imagine Dragons')
    elif message.text == "213":
        user_data[3] = 213
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_213_1.value)
        bot.send_message(message.chat.id, "Оцени предложенных исполнителей оценкой от 1 до 5, где \n"
                                          "1 - тебе совсем не нравится исполнитель \n"
                                          "2 - просто не нравится исполнитель \n"
                                          "3 - не слышал / равнодушен(на) \n"
                                          "4 - нравится исполнитель \n"
                                          "5 - очень нравится исполнитель \n"
                                          "\n"
                                          "Оценивай исполнителей через пробел. Все четыре оценки не должны быть одинаковыми. \n"
                                          "Пример того, что ты должен(на) ввести: \n"
                                          "4 1 2 3")
        bot.send_message(message.chat.id, 'Лана дель Рей , Моргенштерн, Нервы, The Beatles' )
    elif message.text == "210":
        user_data[3] = 210
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_211_1.value)
        bot.send_message(message.chat.id, "Оцени предложенных исполнителей оценкой от 1 до 5, где \n"
                                          "1 - тебе совсем не нравится исполнитель \n"
                                          "2 - просто не нравится исполнитель \n"
                                          "3 - не слышал / равнодушен(на) \n"
                                          "4 - нравится исполнитель \n"
                                          "5 - очень нравится исполнитель \n"
                                          "\n"
                                          "Оценивай исполнителей через пробел. Все четыре оценки не должны быть одинаковыми. \n"
                                          "Пример того, что ты должен(на) ввести: \n"
                                          "4 1 2 3")
        bot.send_message(message.chat.id, 'Sam Smith, Sia, Fall Out Boy, Drake')
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши еще раз")
        return

@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_211_1.value)
def user_estimate211_1(message):
    user_marks = message.text
    if len(user_marks) != 7 or len(set(user_marks)) < 2:
        bot.send_message(message.chat.id, 'Введи четыре оценки через пробел. Все оценки не должны быть одинаковыми')
    else:
        user_data[4] = str(user_marks[0]) + " " + str(user_marks[2]) + " " + str(user_marks[4]) + " " + str(user_marks[6])
        answer, alpha_answer = rec(212, int(user_marks[0]), int(user_marks[2]), int(user_marks[4]), int(user_marks[6]))
        text = "ИСПОЛНИТЕЛЬ \n"
        for i in range(6):
            #text = text + "\n" + str(answer[i]) + "    " + str(round(alpha_answer[i], 2))
            text = text + "\n" + str(answer[i])

        bot.send_message(message.chat.id, text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_GRADE_211_1.value)
        bot.send_message(message.chat.id, 'Довольны ли вы результатом? \n'
                                          'Если да, то напиши ДА \n'
                                          '\n'
                                          'Если нет, то напиши порядок этих исполнителей (Two Feet,	Oh Wonder,	U2,'
                                          '	Kaleo,	Ленинград,	Rag’n’Bone man)'
                                          ' изменяя их ЭТОТ порядок. То есть '
                                          'если ты напишешь 2 3 5 6 1 4 это значит что ты бы поставил(а) на 2-е место'
                                          'Two Feet, на 3-е место Oh Wonder, на 5-е место U2 и т.д.')

@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_212_1.value)
def user_estimate212_1(message):
    user_marks = message.text
    if len(user_marks) != 7 or len(set(user_marks)) < 2:
        bot.send_message(message.chat.id, 'Введи четыре оценки через пробел. Все оценки не должны быть одинаковыми')
    else:
        user_data[4] = str(user_marks[0]) + " " + str(user_marks[2]) + " " + str(user_marks[4]) + " " + str(user_marks[6])
        answer, alpha_answer = rec(213,  int(user_marks[0]), int(user_marks[2]), int(user_marks[4]), int(user_marks[6]))
        text = "ИСПОЛНИТЕЛЬ \n"
        for i in range(6):
            text = text + "\n" + str(answer[i])
        bot.send_message(message.chat.id, text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_GRADE_212_1.value)
        bot.send_message(message.chat.id, 'Довольны ли вы результатом? \n'
                                          'Если да, то напшии ДА \n'
                                          '\n'
                                          'Если нет, то напиши порядок этих исполнителей (Земфира,	Иван Дорн,	'
                                          'INSTASAMKA,	Pink Floyd,	Face,	ЛСП)'
                                          ' изменяя их ЭТОТ порядок. То есть '
                                          'если ты напишешь 2 3 5 6 1 4 это значит что ты бы поставил(а) на 2-е место'
                                          'Земфиру, на 3-е место Ивана Дорна, на 5-е место INSTASAMKA и т.д.')


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_213_1.value)
def user_estimate213_1(message):
    user_marks = message.text
    if len(user_marks) != 7 or len(set(user_marks)) <= 2:
        bot.send_message(message.chat.id, 'Введи четыре оценки через пробел. Все оценки не должны быть одинаковыми')
        return
    else:
        answer, alpha_answer = rec(211,  int(user_marks[0]), int(user_marks[2]), int(user_marks[4]), int(user_marks[6]))
        user_data[4] = str(user_marks[0]) + " " + str(user_marks[2]) + " " + str(user_marks[4]) + " " + str(user_marks[6])
        text = "ИСПОЛНИТЕЛЬ \n"
        for i in range(6):
            text = text + "\n" + str(answer[i])
        bot.send_message(message.chat.id, text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_GRADE_213_1.value)
        bot.send_message(message.chat.id, 'Довольны ли вы результатом? \n'
                                          'Если да, то напшиите ДА \n'
                                          '\n'
                                          'Если нет, то напиши порядок этих исполнителей (Алена Швец,  Рианна,	Эд Ширан,'
                                          '	Клава Кока,	Slava Merlou,	Halsey) изменяя их ЭТОТ порядок. То есть '
                                          'если ты напишешь 2 3 5 6 1 4 это значит что ты бы поставил(а) на 2-е место'
                                          'Алену Швец, на 3-е место Рианну, на 5-е место Эд Ширан и т.д.')


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_GRADE_211_1.value)
def user_estimate_result_211_1(message):
    if message.text.upper() == 'ДА':
        user_data[5] = "YES"
        bot.send_message(message.chat.id, 'Здорово! Тогда оцени следующих исполнителей в таком же формате')
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_211_2.value)
        bot.send_message(message.chat.id, 'Скриптонит, Валерий Меладзе, Кино, Imagine Dragons')

    elif not(len(message.text) != 11 or len(set(message.text))) <=2:
        user_data[5] = ""
        for i in range(6):
            user_data[5] += str(message.text[2*i]) + " "
        bot.send_message(message.chat.id, 'Ок! Мы возьмем это в расчет.\n'
                                          'Оцени тогда следующих исполнителей в том формате, как ты это делал(а) раньше \n'
                                          'Скриптонит, Валерий Меладзе, Кино, Imagine Dragons')
        #print(message.text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_211_2.value)
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши или ДА, или порядок того на какие места ты бы"
                                          " расставил(а) исполнителей через пробелы")
        return


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_GRADE_212_1.value)
def user_estimate_result_212_1(message):
    if message.text.upper() == 'ДА':
        user_data[5] = "YES"
        bot.send_message(message.chat.id, 'Здорово! Тогда оцени следующих исполнителей в таком же формате')
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_212_2.value)
        bot.send_message(message.chat.id, 'Лана дель Рей,	Моргенштерн,	Нервы,	The Beatles')

    elif not(len(message.text) != 11 or len(set(message.text))) <=2:
        user_data[5] = ""
        for i in range(6):
            user_data[5] += str(message.text[2*i]) + " "
        bot.send_message(message.chat.id, 'Ок! Мы мы возьмем это в расчет.\n'
                                          'Оцени тогда следующих исполнителей в том формате, как ты это делал(а) раньше \n'
                                          'Лана дель Рей,	Моргенштерн,	Нервы,	The Beatles')
        #print(message.text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_212_2.value)
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши или ДА, или порядок того на какие места ты бы"
                                          " расставил(а) исполнителей через пробелы")
        return


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_GRADE_213_1.value)
def user_estimate_result_213_1(message):
    if message.text.upper() == 'ДА':
        user_data[5] = "YES"
        bot.send_message(message.chat.id, 'Здорово! Тогда оцени следующих исполнителей в таком же формате')
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_213_2.value)
        bot.send_message(message.chat.id, 'Sam Smith,	Sia,	Fall Out Boy,	Drake')
    elif not(len(message.text) != 11 or len(set(message.text))) <=2:
        user_data[5] = ""
        for i in range(6):
            user_data[5] += str(message.text[2*i]) + " "
        bot.send_message(message.chat.id, 'Ок! Мы мы возьмем это в расчет.\n'
                                          'Оцени тогда следующих исполнителей в том формате, как ты это делал(а) раньше \n'
                                          'Sam Smith,	Sia,	Fall Out Boy,	Drake')
        #print(message.text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_213_2.value)
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши или ДА, или порядок того на какие места ты бы"
                                          " расставил(а) исполнителей через пробелы")
        return


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_211_2.value)
def user_estimate211_2(message):
    user_marks = message.text

    if len(user_marks) != 7 or len(set(user_marks)) < 2:
        bot.send_message(message.chat.id, 'Введи четыре оценки через пробел. Все оценки не должны быть одинаковыми')
        return
    else:
        user_data[6] = str(user_marks[0])+ " "+str(user_marks[2]) + " " + str(user_marks[4]) + " " + str(user_marks[6])
        answer, alpha_answer = rec(213, int(user_marks[0]), int(user_marks[2]), int(user_marks[4]), int(user_marks[6]))
        text = "ИСПОЛНИТЕЛЬ \n"
        for i in range(6):
            text = text + "\n" + str(answer[i])
        bot.send_message(message.chat.id, text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_GRADE_211_2.value)
        bot.send_message(message.chat.id, 'Довольны ли вы результатом? \n'
                                          'Если да, то напшии ДА \n'
                                          '\n'
                                          'Если нет, то напиши порядок этих исполнителей (Земфира,	Иван Дорн,	'
                                          'INSTASAMKA,	Pink Floyd,	Face,	ЛСП)'
                                          ' изменяя их ЭТОТ порядок. То есть '
                                          'если ты напишешь 2 3 5 6 1 4 это значит что ты бы поставил(а) на 2-е место'
                                          'Земфиру, на 3-е место Ивана Дорна, на 5-е место INSTASAMKA и т.д.')
@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_212_2.value)
def user_estimate212_2(message):
    user_marks = message.text
    if len(user_marks) != 7 or len(set(user_marks)) < 2:
        bot.send_message(message.chat.id, 'Введи четыре оценки через пробел. Все оценки не должны быть одинаковыми')
        return
    else:
        user_data[6] = str(user_marks[0])+ " "+str(user_marks[2]) + " " + str(user_marks[4]) + " " + str(user_marks[6])
        answer, alpha_answer = rec(211,  int(user_marks[0]), int(user_marks[2]), int(user_marks[4]), int(user_marks[6]))
        text = "ИСПОЛНИТЕЛЬ \n"
        for i in range(6):
            text = text + "\n" + str(answer[i])
        bot.send_message(message.chat.id, text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_GRADE_212_2.value)
        bot.send_message(message.chat.id, 'Довольны ли вы результатом? \n'
                                          'Если да, то напшии ДА \n'
                                          '\n'
                                          'Если нет, то напиши порядок этих исполнителей (Алена Швец,  Рианна,	Эд Ширан,'
                                          '	Клава Кока,	Slava Merlou,	Halsey) изменяя их ВОТ ЭТОТ порядок. То есть '
                                          'если ты напишешь 2 3 5 6 1 4 это значит что ты бы поставил(а) на 2-е место'
                                          'Алену Швец, на 3-е место Рианну, на 5-е место Эд Ширан и т.д.')



@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_213_2.value)
def user_estimate213_2(message):
    user_marks = message.text
    if len(user_marks) != 7 or len(set(user_marks)) <= 2:
        bot.send_message(message.chat.id, 'Введи четыре оценки через пробел. Все оценки не должны быть одинаковыми')
        return
    else:
        user_data[6] = str(user_marks[0])+ " "+str(user_marks[2]) + " " + str(user_marks[4]) + " " + str(user_marks[6])
        answer, alpha_answer = rec(212,  int(user_marks[0]), int(user_marks[2]), int(user_marks[4]), int(user_marks[6]))
        text = "ИСПОЛНИТЕЛЬ \n"
        for i in range(6):
            text = text + "\n" + str(answer[i])
        bot.send_message(message.chat.id, text)
        dbworker.set_state(message.chat.id, config.States.S_ESTIMATE_GRADE_213_2.value)
        bot.send_message(message.chat.id, 'Довольны ли вы результатом? \n'
                                          'Если да, то напшии ДА \n'
                                          'Если нет, то напиши порядок этих исполнителей (Two Feet,	Oh Wonder,	U2,'
                                          '	Kaleo,	Ленинград,	Rag’n’Bone man)'
                                          ' изменяя их ЭТОТ порядок. То есть '
                                          'если ты напишешь 2 3 5 6 1 4 это значит что ты бы поставил(а) на 2-е место'
                                          'Two Feet, на 3-е место Oh Wonder, на 5-е место U2 и т.д.')

@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_GRADE_211_2.value)
def user_estimate_result_211_2(message):
    if message.text.upper() == 'ДА':
        user_data[7] = "YES"
        bot.send_message(message.chat.id, 'Здорово! Спасибо за работу с ботом!')
        filename = 'text.xlsx'
        wb = load_workbook(filename)
        file = wb.active
        file.append(user_data)
        wb.save(filename=filename)
        dbworker.set_state(message.chat.id, config.States.S_START.value)
    elif not(len(message.text) != 11 or len(set(message.text))) <=2:
        user_data[7] = ""
        for i in range(6):
            user_data[7] += str(message.text[2*i]) + " "
        bot.send_message(message.chat.id, 'Ок! Мы возьмем это в расчет.\n'
                                          'Спасибо за работу с ботом!')
        filename = 'test.xlsx'
        wb = load_workbook(filename)
        file = wb.active
        file.append(user_data)
        wb.save(filename=filename)
        dbworker.set_state(message.chat.id, config.States.S_START.value)
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши или ДА, или порядок того на какие места ты бы"
                                          " расставил(а) исполнителей через пробелы")
        return

@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) == config.States.S_ESTIMATE_GRADE_212_2.value)
def user_estimate_result_212_2(message):
    if message.text.upper() == 'ДА':
        user_data[7] = "YES"
        bot.send_message(message.chat.id, 'Здорово! Спасибо за работу с ботом!')
        filename = 'test.xlsx'
        wb = load_workbook(filename)
        file = wb.active
        file.append(user_data)
        wb.save(filename=filename)
        dbworker.set_state(message.chat.id, config.States.S_START.value)
    elif not(len(message.text) != 11 or len(set(message.text))) <=2:
        user_data[7] = ""
        for i in range(6):
            user_data[7] += str(message.text[2*i]) + " "
        bot.send_message(message.chat.id, 'Ок! Мы мы возьмем это в расчет.\n'
                                          'Спасибо за работу с ботом!')
        filename = 'test.xlsx'
        wb = load_workbook(filename)
        file = wb.active
        file.append(user_data)
        wb.save(filename=filename)
        dbworker.set_state(message.chat.id, config.States.S_START.value)
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши или ДА, или порядок того на какие места ты бы"
                                          " расставил(а) исполнителей через пробелы")
        return


@bot.message_handler(func = lambda message: dbworker.get_current_state(message.chat.id) ==
                                            config.States.S_ESTIMATE_GRADE_213_2.value)
def user_estimate_result_213_2(message):
    if message.text.upper() == 'ДА':
        user_data[7] = "YES"
        bot.send_message(message.chat.id, 'Здорово! Спасибо за работу с ботом!')

        filename = 'test.xlsx'
        wb = load_workbook(filename)
        file = wb.active
        file.append(user_data)
        wb.save(filename=filename)

        dbworker.set_state(message.chat.id, config.States.S_START.value)
    elif not(len(message.text) != 11 or len(set(message.text))) <=2:
        user_data[7] = ""
        for i in range(6):
            user_data[7] += str(message.text[2*i]) + " "
        bot.send_message(message.chat.id, 'Ок! Мы мы возьмем это в расчет.\n'
                                          'Спасибо за работу с ботом!')
        filename = 'test.xlsx'
        wb = load_workbook(filename)
        file = wb.active
        file.append(user_data)
        wb.save(filename=filename)
        dbworker.set_state(message.chat.id, config.States.S_START.value)
    else:
        bot.send_message(message.chat.id, "Я тебя не понял. Напиши или ДА, или порядок того на какие места ты бы"
                                          " расставил(а) исполнителей через пробелы")
        return


bot.infinity_polling()